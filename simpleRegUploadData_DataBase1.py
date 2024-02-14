
# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
from datetime import datetime
import pymysql  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
#wb = load_workbook('//home//hakeem//Desktop//excel.xlsx')
wb = load_workbook('excel.xlsx') 
  
#database connection
connection = pymysql.connect(host="localhost",user="root",passwd="",database="Student")
cursor = connection.cursor()

# create the sheet object 
sheet = wb.active 
today=datetime.now()  
  
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 10
  
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Date"
    sheet.cell(row=1, column=2).value = "Form Number"
    sheet.cell(row=1, column=3).value = "Name"
    sheet.cell(row=1, column=4).value = "Course"
    sheet.cell(row=1, column=5).value = "Phone Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"
    sheet.cell(row=1, column=8).value = "Status"
  
  
# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the course_field box 
    course_field.focus_set() 
  
  
# Function to set focus 
def focus2(event): 
    # set focus on the status_field box 
    status_field.focus_set() 
  
  
# Function to set focus 
def focus3(event): 
    # set focus on the form_no_field box 
    form_no_field.focus_set() 
  
  
# Function to set focus 
def focus4(event): 
    # set focus on the contact_no_field box 
    contact_no_field.focus_set() 
  
  
# Function to set focus 
def focus5(event): 
    # set focus on the email_id_field box 
    email_id_field.focus_set() 
  
  
# Function to set focus 
def focus6(event): 
    # set focus on the address_field box 
    address_field.focus_set() 
# Function to set focus 
def focus7(event):
    # set focus on the address_field box 
    roll_number_field.focus_set()  
  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    name_field.delete(0, END) 
    course_field.delete(0, END) 
    status_field.delete(0, END) 
    form_no_field.delete(0, END) 
    contact_no_field.delete(0, END) 
    email_id_field.delete(0, END) 
    address_field.delete(0, END) 
    roll_number_field.delete(0, END) 
  
def sub_option():
    print (regDetails.get())
    if regDetails.get() == 1:
    	temp = IntVar()    
    	Checkbutton(root, text="Demo", variable=temp).grid(row=11,column=1, sticky=N)
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
    print (regDetails.get())      
    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" and
        course_field.get() == "" and
        status_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == "" and 
        roll_number_field.get() == ""): 
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = today.strftime("%d-%m-%Y %H:%M") 
        sheet.cell(row=current_row + 1, column=2).value = form_no_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = name_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = course_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get() 
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get() 
        sheet.cell(row=current_row + 1, column=7).value = address_field.get() 
        sheet.cell(row=current_row + 1, column=8).value = status_field.get() 
	# Data inserting into database Start
        if regDetails.get() == 1 and examDetails.get() == 1:
                insert_reg_record = "INSERT INTO Registration(FORM_NUMBER,ROLL_NUMBER,NAME,COURSE,SEMISTER,CONTACT_NUMBER,EMAIL_ID,ADDRESS) VALUES('%s','%s','%s','%s','%s','%s','%s','%s');"%(form_no_field.get(),roll_number_field.get(),name_field.get(),course_field.get(),status_field.get(),contact_no_field.get(),email_id_field.get(),address_field.get())
                insert_exam_record = "INSERT INTO Exam(FORM_NUMBER,ROLL_NUMBER,NAME,COURSE,SEMISTER,CONTACT_NUMBER,EMAIL_ID,ADDRESS) VALUES('%s','%s','%s','%s','%s','%s','%s','%s');"%(form_no_field.get(),roll_number_field.get(),name_field.get(),course_field.get(),status_field.get(),contact_no_field.get(),email_id_field.get(),address_field.get())
        elif regDetails.get() == 1:
                insert_reg_record = "INSERT INTO Registration(FORM_NUMBER,ROLL_NUMBER,NAME,COURSE,SEMISTER,CONTACT_NUMBER,EMAIL_ID,ADDRESS) VALUES('%s','%s','%s','%s','%s','%s','%s','%s');"%(form_no_field.get(),roll_number_field.get(),name_field.get(),course_field.get(),status_field.get(),contact_no_field.get(),email_id_field.get(),address_field.get())
        elif examDetails.get() == 1:
                insert_exam_record = "INSERT INTO Exam(FORM_NUMBER,ROLL_NUMBER,NAME,COURSE,SEMISTER,CONTACT_NUMBER,EMAIL_ID,ADDRESS) VALUES('%s','%s','%s','%s','%s','%s','%s','%s');"%(form_no_field.get(),roll_number_field.get(),name_field.get(),course_field.get(),status_field.get(),contact_no_field.get(),email_id_field.get(),address_field.get())
        else :
                print ("Invalid Details")
        if regDetails.get() == 1:
           cursor.execute(insert_reg_record)
        if examDetails.get() == 1:
           cursor.execute(insert_exam_record)
        if regDetails.get() == 1 or examDetails.get() == 1:
           connection.commit()
	# Data inserting into database End 
  
        # save the file 
        wb.save('//home//hakeem//Desktop//excel.xlsx') 
  
        # set focus on the name_field box 
        name_field.focus_set() 
  
        # call the clear() function 
        clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light green') 
  
    # set the title of GUI window 
    root.title("registration form") 
  
    # set the configuration of GUI window 
    root.geometry("500x300") 
  
    excel() 
  
    # create a Form label 
    heading = Label(root, text="Form", bg="light green") 
  
    # create a Name label 
    name = Label(root, text="Name", bg="light green") 
  
    # create a Course label 
    course = Label(root, text="Course", bg="light green") 
  
    # create a Semester label 
    status = Label(root, text="Status", bg="light green") 
  
    # create a Form No. lable 
    form_no = Label(root, text="Form No.", bg="light green") 
  
    # create a Contact No. label 
    contact_no = Label(root, text="Contact No.", bg="light green") 
  
    # create a Email id label 
    email_id = Label(root, text="Email id", bg="light green") 
  
    # create a address label 
    address = Label(root, text="Address", bg="light green") 
    # create a Roll Number label 
    roll_number = Label(root, text="Roll Number", bg="light green") 
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    name.grid(row=1, column=0) 
    course.grid(row=2, column=0) 
    status.grid(row=3, column=0) 
    form_no.grid(row=4, column=0) 
    contact_no.grid(row=5, column=0) 
    email_id.grid(row=6, column=0) 
    address.grid(row=7, column=0) 
    roll_number.grid(row=8, column=0) 
  
    # create a text entry box 
    # for typing the information 
    name_field = Entry(root) 
    course_field = Entry(root) 
    status_field = Entry(root) 
    form_no_field = Entry(root) 
    contact_no_field = Entry(root) 
    email_id_field = Entry(root) 
    address_field = Entry(root) 
    roll_number_field = Entry(root) 
  
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    name_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    course_field.bind("<Return>", focus2) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
    status_field.bind("<Return>", focus3) 
  
    # whenever the enter key is pressed 
    # then call the focus4 function 
    form_no_field.bind("<Return>", focus4) 
  
    # whenever the enter key is pressed 
    # then call the focus5 function 
    contact_no_field.bind("<Return>", focus5) 
  
    # whenever the enter key is pressed 
    # then call the focus6 function 
    email_id_field.bind("<Return>", focus6) 
    roll_number_field.bind("<Return>", focus7) 
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    name_field.grid(row=1, column=1, ipadx="100") 
    course_field.grid(row=2, column=1, ipadx="100") 
    status_field.grid(row=3, column=1, ipadx="100") 
    form_no_field.grid(row=4, column=1, ipadx="100") 
    contact_no_field.grid(row=5, column=1, ipadx="100") 
    email_id_field.grid(row=6, column=1, ipadx="100") 
    address_field.grid(row=7, column=1, ipadx="100") 
    roll_number_field.grid(row=8, column=1, ipadx="100") 
  
    # call excel function 
    excel() 
    regDetails = IntVar()
    #Checkbutton(root, text="Registration", variable=regDetails).grid(row=9,column=1, sticky=N)
    Checkbutton(root, text="Registration", variable=regDetails,command=sub_option).grid(row=9,column=1, sticky=N)
    examDetails = IntVar()
    Checkbutton(root, text="Enquiry", variable=examDetails).grid(row=9,column=1, sticky=W)  

    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Green", command=insert) 
    submit.grid(row=10, column=1,sticky=W) 
  
    #Quit Button creating
    quite = Button(root, text="Quite", fg="Black",
                            bg="Red", command=root.destroy)
    quite.grid(row=10, column=1,sticky=N)

    # start the GUI 
    root.mainloop() 

